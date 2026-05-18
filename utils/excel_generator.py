import io
import pandas as pd

def generate_excel_report(sessions, roadmap_title="Combined"):
    """
    Generate a styled, downloadable Excel report of the user's focus sessions.
    """
    df = pd.DataFrame(sessions)
    if df.empty:
        return None
    
    # Select and rename columns for a clean user report
    rename_cols = {
        'start_time': 'Start Time',
        'end_time': 'End Time',
        'planned_duration': 'Planned Duration (mins)',
        'actual_duration': 'Actual Duration (mins)',
        'focus_score': 'Focus Score (%)',
        'distraction_count': 'Total Distractions',
        'phone_count': 'Phone Distractions',
        'drowsy_count': 'Drowsiness Distractions',
        'zone_out_count': 'Zone-out Distractions',
        'pause_count': 'Pauses',
        'notes': 'Notes'
    }
    
    # Filter to only existing columns in DataFrame
    available_cols = [col for col in rename_cols.keys() if col in df.columns]
    report_df = df[available_cols].rename(columns=rename_cols)
    
    # Sort by Start Time descending (most recent first)
    if 'Start Time' in report_df.columns:
        report_df = report_df.sort_values(by='Start Time', ascending=False)
    
    # Write to a bytes buffer
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        report_df.to_excel(writer, index=False, sheet_name='Focus & Distractions Report')
        
        workbook = writer.book
        worksheet = writer.sheets['Focus & Distractions Report']
        
        # Style sheet to make it feel premium
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Color definitions
        primary_color = "6366F1"  # Premium Indigo
        border_color = "CBD5E1"   # Slate 300
        
        # Fonts
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        regular_font = Font(name="Calibri", size=10)
        bold_font = Font(name="Calibri", size=10, bold=True)
        
        # Fills
        header_fill = PatternFill(start_color=primary_color, end_color=primary_color, fill_type="solid")
        zebra_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        
        # Alignments
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        
        # Borders
        thin_border = Border(
            left=Side(style='thin', color=border_color),
            right=Side(style='thin', color=border_color),
            top=Side(style='thin', color=border_color),
            bottom=Side(style='thin', color=border_color)
        )
        
        # Format headers
        for col_idx in range(1, len(report_df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            
        # Format data rows
        for row_idx in range(2, len(report_df) + 2):
            is_even = (row_idx % 2 == 0)
            for col_idx in range(1, len(report_df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = regular_font
                cell.border = thin_border
                
                # Zebra striping
                if is_even:
                    cell.fill = zebra_fill
                
                # Check column name for custom formatting
                col_name = report_df.columns[col_idx - 1]
                
                # Focus Score color coding
                if col_name == 'Focus Score (%)':
                    cell.font = bold_font
                    try:
                        score_val = float(cell.value)
                        if score_val >= 80:
                            cell.font = Font(name="Calibri", size=10, bold=True, color="10B981") # Green
                        elif score_val >= 60:
                            cell.font = Font(name="Calibri", size=10, bold=True, color="F59E0B") # Yellow
                        else:
                            cell.font = Font(name="Calibri", size=10, bold=True, color="EF4444") # Red
                    except:
                        pass
                
                # Text alignment
                if col_name in ['Notes', 'Start Time', 'End Time']:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
        
        # Adjust column widths dynamically
        for col in worksheet.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if '\n' in val:
                    val = max(val.split('\n'), key=len)
                max_len = max(max_len, len(val))
            col_letter = col[0].column_letter
            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    buffer.seek(0)
    return buffer
